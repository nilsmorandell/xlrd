"""
Cashflow-Berechnung Excel Generator
Erstellt eine Excel-Tabelle zur Berechnung des Cashflows
basierend auf der indirekten Methode.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# Farben
BLACK_FILL = "1F1F1F"
GREEN_FILL = "2E7D32"
LIGHT_GREEN_FILL = "C8E6C9"
HEADER_FONT_COLOR = "FFFFFF"
LABEL_FONT_COLOR = "1F1F1F"

THIN = Side(style="thin", color="AAAAAA")
THICK = Side(style="medium", color="333333")
NO_BORDER = Side(style=None)


def make_border(top=None, bottom=None, left=None, right=None):
    return Border(
        top=top or NO_BORDER,
        bottom=bottom or NO_BORDER,
        left=left or NO_BORDER,
        right=right or NO_BORDER,
    )


def style_header(cell, text, bg_color=BLACK_FILL, font_color=HEADER_FONT_COLOR, bold=True, size=11):
    cell.value = text
    cell.font = Font(name="Calibri", bold=bold, color=font_color, size=size)
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_result(cell, text, bg_color=GREEN_FILL, font_color=HEADER_FONT_COLOR, bold=True):
    cell.value = text
    cell.font = Font(name="Calibri", bold=bold, color=font_color, size=11)
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_input_label(cell, text, indent=0):
    cell.value = text
    cell.font = Font(name="Calibri", size=10, color=LABEL_FONT_COLOR)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                                indent=indent, wrap_text=True)
    cell.fill = PatternFill("solid", fgColor="F5F5F5")


def style_input_value(cell, value=0):
    cell.value = value
    cell.font = Font(name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '#,##0.00 €'
    cell.fill = PatternFill("solid", fgColor="FFFFFF")


def style_formula_cell(cell, formula):
    cell.value = formula
    cell.font = Font(name="Calibri", size=10, bold=True)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '#,##0.00 €'
    cell.fill = PatternFill("solid", fgColor="E8F5E9")


def style_total_formula(cell, formula, bg_color=GREEN_FILL):
    cell.value = formula
    cell.font = Font(name="Calibri", size=11, bold=True, color=HEADER_FONT_COLOR)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '#,##0.00 €'
    cell.fill = PatternFill("solid", fgColor=bg_color)


def create_cashflow_excel(filename="Cashflow_Berechnung.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cashflow-Berechnung"

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 2   # linker Rand
    ws.column_dimensions["B"].width = 52  # Beschriftung
    ws.column_dimensions["C"].width = 5   # Vorzeichen
    ws.column_dimensions["D"].width = 18  # Betrag
    ws.column_dimensions["E"].width = 2   # Abstand
    ws.column_dimensions["F"].width = 30  # Rechte Box Beschriftung
    ws.column_dimensions["G"].width = 18  # Rechte Box Betrag

    # Zeilenhöhen
    row_heights = {
        1: 10, 2: 28, 3: 10, 4: 22, 5: 22, 6: 22, 7: 28,
        8: 10, 9: 22, 10: 22, 11: 28, 12: 10, 13: 22, 14: 22,
        15: 28, 16: 10, 17: 22, 18: 22, 19: 22, 20: 22, 21: 28,
        22: 10, 23: 28, 24: 28, 25: 28,
    }
    for r, h in row_heights.items():
        ws.row_dimensions[r].height = h

    # ===== TITEL =====
    ws.merge_cells("B2:D2")
    cell = ws["B2"]
    style_header(cell, "Cash Flow – Berechnung", bg_color=BLACK_FILL, size=14)

    # ===== BLOCK 1: Traditioneller Cashflow =====
    ws.merge_cells("B4:C4")
    style_input_label(ws["B4"], "Ergebnis aus G&V (Jahresüberschuss/-fehlbetrag)")
    style_input_value(ws["D4"])

    ws.merge_cells("B5:C5")
    style_input_label(ws["B5"], "+/–  Abschreibung / Aufwertung")
    style_input_value(ws["D5"])

    ws.merge_cells("B6:C6")
    style_input_label(ws["B6"], "+/–  Erhöhung / Verminderung Rückstellungen")
    style_input_value(ws["D6"])

    ws.merge_cells("B7:C7")
    style_result(ws["B7"], "1.  = Traditioneller Cashflow (Brutto Cashflow)")
    style_total_formula(ws["D7"], "=D4+D5+D6")

    # ===== BLOCK 2: Operativer Cashflow =====
    ws.merge_cells("B9:C9")
    style_input_label(ws["B9"], "+/–  Verminderung / Erhöhung der Forderungen, Vorräte etc.")
    style_input_value(ws["D9"])

    ws.merge_cells("B10:C10")
    style_input_label(ws["B10"], "+/–  Erhöhung / Verminderung der Lieferverbindlichkeiten etc.")
    style_input_value(ws["D10"])

    ws.merge_cells("B11:C11")
    style_result(ws["B11"], "2.  = Cash Flow aus Geschäftstätigkeit (operativer Cashflow)")
    style_total_formula(ws["D11"], "=D7+D9+D10")

    # ===== BLOCK 3: Investitionstätigkeit =====
    ws.merge_cells("B13:C13")
    style_input_label(ws["B13"], "+    Einzahlungen aus Anlageabgängen")
    style_input_value(ws["D13"])

    ws.merge_cells("B14:C14")
    style_input_label(ws["B14"], "–    Auszahlungen für Anlageinvestitionen")
    style_input_value(ws["D14"])

    ws.merge_cells("B15:C15")
    style_result(ws["B15"], "3.  = Cash Flow aus Investitionstätigkeit")
    style_total_formula(ws["D15"], "=D13+D14")

    # ===== BLOCK 4: Finanzierungstätigkeit =====
    ws.merge_cells("B17:C17")
    style_input_label(ws["B17"], "+    Einzahlungen aus Zuführungen von Eigenkapital")
    style_input_value(ws["D17"])

    ws.merge_cells("B18:C18")
    style_input_label(ws["B18"], "–    Auszahlungen an die Eigentümer")
    style_input_value(ws["D18"])

    ws.merge_cells("B19:C19")
    style_input_label(ws["B19"], "+    Einzahlungen aus Aufnahme von Finanzverbindlichkeiten")
    style_input_value(ws["D19"])

    ws.merge_cells("B20:C20")
    style_input_label(ws["B20"], "–    Auszahlungen aus Rückzahlung von Finanzverbindlichkeiten")
    style_input_value(ws["D20"])

    ws.merge_cells("B21:C21")
    style_result(ws["B21"], "4.  = Cash Flow aus Finanzierungstätigkeit")
    style_total_formula(ws["D21"], "=D17+D18+D19+D20")

    # ===== RECHTE BOX: Gesamtübersicht =====
    style_header(ws["F23"], "Gesamter Zahlungsmittelzufluss\n(Positionen 1 bis 4)",
                 bg_color=BLACK_FILL, size=10)
    style_formula_cell(ws["G23"], "=D11+D15+D21")

    style_input_label(ws["F24"], "+   Anfangsbestand an Finanzmitteln\n    (zu Jahresbeginn)", indent=1)
    style_input_value(ws["G24"])

    style_result(ws["F25"], "= Finanzmittelbestand am Ende\n  des Geschäftsjahres", bg_color=GREEN_FILL)
    style_total_formula(ws["G25"], "=G23+G24", bg_color=GREEN_FILL)

    # Rahmen für rechte Box
    for row in [23, 24, 25]:
        for col in ["F", "G"]:
            cell = ws[f"{col}{row}"]
            cell.border = Border(
                top=Side(style="thin", color="888888"),
                bottom=Side(style="thin", color="888888"),
                left=Side(style="thin", color="888888"),
                right=Side(style="thin", color="888888"),
            )

    # Rahmen für Eingabefelder
    for row in [4, 5, 6, 9, 10, 13, 14, 17, 18, 19, 20]:
        ws[f"D{row}"].border = Border(
            bottom=Side(style="thin", color="AAAAAA"),
            left=Side(style="thin", color="AAAAAA"),
            right=Side(style="thin", color="AAAAAA"),
        )

    # Zellhöhe für rechte Box anpassen
    ws.row_dimensions[23].height = 36
    ws.row_dimensions[24].height = 36
    ws.row_dimensions[25].height = 36

    # Rechte Box Textformatierung
    ws["F23"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws["F24"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws["F25"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws["G23"].alignment = Alignment(horizontal="right", vertical="center")
    ws["G23"].font = Font(name="Calibri", size=10, bold=True)
    ws["G23"].number_format = '#,##0.00 €'
    ws["G23"].fill = PatternFill("solid", fgColor="E8F5E9")

    # Druckbereich
    ws.print_area = "A1:G26"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True

    wb.save(filename)
    print(f"Excel-Datei '{filename}' wurde erfolgreich erstellt.")
    return filename


if __name__ == "__main__":
    create_cashflow_excel()
